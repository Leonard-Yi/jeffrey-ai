# D:/Epstein.AI/scripts/research/write_excel.py
"""
将公司数据写入横向布局的 Excel 文件。

布局:
            | 公司A  | 公司B  | 公司C  |
成立时间    |  xxx   |  xxx   |  xxx   |
所在地      |  xxx   |  xxx   |  xxx   |
...

文件命名: 行业名_公司研究_YYYYMMDD.xlsx
策略: 已存在的数据覆盖写入；新增公司追加列；缺失字段留空。
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
import datetime
from scripts.research.config import OUTPUT_DIR
from scripts.research.constants import ALL_FIELDS


Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)


def get_filename(industry: str) -> Path:
    """生成文件名：行业名_公司研究_YYYYMMDD.xlsx"""
    date = datetime.date.today().strftime("%Y%m%d")
    safe_industry = industry.replace("/", "_").replace("\\", "_")
    return Path(OUTPUT_DIR) / f"{safe_industry}_公司研究_{date}.xlsx"


def init_workbook(filepath: Path, companies: list[str]) -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet, list[str]]:
    """
    创建或打开 Excel 文件，初始化表头。
    返回 (workbook, worksheet, all_companies) — all_companies 包含已有的 + 新公司。
    """
    if filepath.exists():
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        # 读取已有公司名
        existing = []
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                existing.append(str(val))
        # 追加新公司
        all_companies = existing.copy()
        for company in companies:
            if company not in all_companies:
                all_companies.append(company)
                # 写新表头
                ws.cell(row=1, column=len(all_companies) + 1, value=company)
        # 确保所有字段行存在
        for row, field in enumerate(ALL_FIELDS, start=2):
            if ws.cell(row=row, column=1).value is None:
                ws.cell(row=row, column=1, value=field)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "公司研究"
        ws.cell(row=1, column=1, value="字段")
        for col, company in enumerate(companies, start=2):
            ws.cell(row=1, column=col, value=company)
        for row, field in enumerate(ALL_FIELDS, start=2):
            ws.cell(row=row, column=1, value=field)
        all_companies = companies

    return wb, ws, all_companies


def write_company_data(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    company: str,
    data: dict,
    all_companies: list[str],
):
    """
    将公司数据写入指定列。覆盖模式：先清空该列再写入。
    如果公司不在 all_companies 中，先追加列再写入。
    """
    if company not in all_companies:
        all_companies.append(company)
        ws.cell(row=1, column=len(all_companies) + 1, value=company)

    col_index = all_companies.index(company) + 2  # +2 因为第1列是字段名，第2列开始是公司

    for row, field in enumerate(ALL_FIELDS, start=2):
        cell = ws.cell(row=row, column=col_index)
        value = data.get(field, "")
        if value is None:
            value = ""
        cell.value = value


def save_workbook(wb: openpyxl.Workbook, filepath: Path):
    """保存 workbook。"""
    try:
        import fcntl
        lock_file = filepath.with_suffix(filepath.suffix + ".lock")
        with open(lock_file, "w") as lf:
            fcntl.flock(lf.fileno(), fcntl.LOCK_EX)
            wb.save(filepath)
            fcntl.flock(lf.fileno(), fcntl.LOCK_UN)
    except (ImportError, AttributeError):
        # Windows 或没有 fcntl 的环境，直接保存
        wb.save(filepath)
    print(f"[OK] 已保存: {filepath.name}")


if __name__ == "__main__":
    # CLI 测试
    import json
    import sys

    if len(sys.argv) < 4:
        print("用法: python write_excel.py <行业名> <公司名> <数据JSON|JSON文件路径>")
        sys.exit(1)

    industry = sys.argv[1]
    company = sys.argv[2]
    data_arg = sys.argv[3]

    if data_arg.startswith("{"):
        data = json.loads(data_arg)
    else:
        with open(data_arg, encoding="utf-8") as f:
            data = json.load(f)

    companies = [company]
    filepath = get_filename(industry)
    wb, ws, all_companies = init_workbook(filepath, companies)
    write_company_data(ws, company, data, all_companies)
    save_workbook(wb, filepath)